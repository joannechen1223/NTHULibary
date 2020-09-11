import requests
from openpyxl import load_workbook


class PdfService(object):

    _TIMEOUT = 15

    def download_pdfs(self, input_path, output_path):
        wb = load_workbook(filename=input_path)
        ws = wb.active
        print(input_path, ws.max_row)
        for row in range(2, ws.max_row+1):
            cell = ws.cell(row=row, column=8)
            if cell.value == "pdf":
                filename = ws.cell(row=row, column=4). \
                    value.replace(" ", "_") + ".pdf"
                self._save_pdf(
                    cell.hyperlink.target,
                    output_path + "/" + filename
                )

    def _save_pdf(self, url, output_path):
        print(url)
        print(output_path)
        try:
            res = requests.get(url, timeout=self._TIMEOUT)
            with open(output_path, 'wb') as f:
                f.write(res.content)
        except Exception:
            print("timeout")
